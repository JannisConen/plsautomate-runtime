"""Generate test fixture files for the ZIP + Excel mapping scenario.

Scenario: Incoming orders with attached documents (ZIP subfolders = order numbers)
joined to SAP-style VBAK (order headers) and VBAP (order line items) Excel sheets.

Run from this directory:
    python generate.py
"""

from __future__ import annotations

import io
import zipfile
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

HERE = Path(__file__).parent


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _fake_pdf(title: str) -> bytes:
    """Minimal valid-looking PDF bytes (not a real PDF, just plausible content)."""
    return (
        f"%PDF-1.4\n%Test fixture\n1 0 obj\n<< /Type /Catalog >>\nendobj\n"
        f"%%EOF\n% {title}\n"
    ).encode()


def _style_header_row(ws, n_cols: int, fill_color: str = "1F4E79") -> None:
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


# ─── Order data ───────────────────────────────────────────────────────────────

ORDERS = [
    {
        "VBELN": "0000012345",
        "ERDAT": "2024-03-01",
        "KUNNR": "0000001001",
        "NAME1": "Müller Maschinenbau GmbH",
        "NETWR": 15200.00,
        "WAERK": "EUR",
        "VKORG": "1000",
        "VTWEG": "10",
        "SPART": "00",
        "files": [
            ("Angebot_12345.pdf", "Angebot / Offer"),
            ("Spezifikation_12345.pdf", "Technische Spezifikation"),
        ],
    },
    {
        "VBELN": "0000012346",
        "ERDAT": "2024-03-03",
        "KUNNR": "0000001002",
        "NAME1": "Schmidt & Partner AG",
        "NETWR": 8750.50,
        "WAERK": "EUR",
        "VKORG": "1000",
        "VTWEG": "10",
        "SPART": "00",
        "files": [
            ("Bestellung_12346.pdf", "Kundenbestellung"),
            ("Rahmenvertrag_12346.pdf", "Rahmenvertrag 2024"),
            ("Zeichnung_12346.pdf", "Technische Zeichnung Rev. B"),
        ],
    },
    {
        "VBELN": "0000012347",
        "ERDAT": "2024-03-05",
        "KUNNR": "0000001003",
        "NAME1": "Technik Solutions KG",
        "NETWR": 31800.00,
        "WAERK": "EUR",
        "VKORG": "1000",
        "VTWEG": "10",
        "SPART": "00",
        "files": [
            ("Auftrag_12347.pdf", "Kundenauftrag"),
            ("Lieferschein_12347.pdf", "Lieferschein"),
        ],
    },
    {
        "VBELN": "0000012348",
        "ERDAT": "2024-03-07",
        "KUNNR": "0000001004",
        "NAME1": "Becker Automotive GmbH",
        "NETWR": 4500.00,
        "WAERK": "EUR",
        "VKORG": "1000",
        "VTWEG": "20",
        "SPART": "10",
        "files": [
            ("Order_12348.pdf", "Purchase Order"),
        ],
    },
]

POSITIONS = [
    # Order 12345
    {"VBELN": "0000012345", "POSNR": "000010", "MATNR": "MAT-STAHL-001", "ARKTX": "Stahlplatte 10mm S235",       "KWMENG": 50.0,  "VRKME": "ST",  "NETPR": 185.00, "MWSBP": 19.0},
    {"VBELN": "0000012345", "POSNR": "000020", "MATNR": "MAT-SCHRAUBE-M8", "ARKTX": "Schraube M8x20 DIN 933",   "KWMENG": 500.0, "VRKME": "ST",  "NETPR": 0.48,   "MWSBP": 19.0},
    {"VBELN": "0000012345", "POSNR": "000030", "MATNR": "MAT-DICHT-001",  "ARKTX": "Dichtung EPDM 50x50",       "KWMENG": 100.0, "VRKME": "ST",  "NETPR": 1.20,   "MWSBP": 19.0},
    # Order 12346
    {"VBELN": "0000012346", "POSNR": "000010", "MATNR": "MAT-ALU-001",    "ARKTX": "Aluminiumprofil 40x40",      "KWMENG": 30.0,  "VRKME": "M",   "NETPR": 8.50,   "MWSBP": 19.0},
    {"VBELN": "0000012346", "POSNR": "000020", "MATNR": "MAT-GUMMI-002",  "ARKTX": "Gummidichtung rund Ø80",    "KWMENG": 100.0, "VRKME": "ST",  "NETPR": 45.00,  "MWSBP": 19.0},
    # Order 12347
    {"VBELN": "0000012347", "POSNR": "000010", "MATNR": "MAT-STAHL-001",  "ARKTX": "Stahlplatte 10mm S235",      "KWMENG": 200.0, "VRKME": "ST",  "NETPR": 178.00, "MWSBP": 19.0},
    {"VBELN": "0000012347", "POSNR": "000020", "MATNR": "MAT-BOLZEN-SET", "ARKTX": "Bolzensatz komplett",        "KWMENG": 20.0,  "VRKME": "SET", "NETPR": 145.00, "MWSBP": 19.0},
    {"VBELN": "0000012347", "POSNR": "000030", "MATNR": "MAT-LACK-9005",  "ARKTX": "Lack RAL 9005 tiefschwarz", "KWMENG": 10.0,  "VRKME": "KG",  "NETPR": 78.00,  "MWSBP": 19.0},
    # Order 12348
    {"VBELN": "0000012348", "POSNR": "000010", "MATNR": "MAT-KFZ-HALTER", "ARKTX": "KFZ-Halterung Typ A",       "KWMENG": 30.0,  "VRKME": "ST",  "NETPR": 150.00, "MWSBP": 19.0},
]


# ─── VBAK.xlsx ────────────────────────────────────────────────────────────────

def create_vbak(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VBAK"

    headers = ["VBELN", "ERDAT", "KUNNR", "NAME1", "NETWR", "WAERK", "VKORG", "VTWEG", "SPART"]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for o in ORDERS:
        ws.append([o[h] for h in headers])

    # Column widths
    for col, width in zip("ABCDEFGHI", [15, 12, 14, 30, 12, 8, 8, 8, 8]):
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"  Created {path.name}")


# ─── VBAP.xlsx ────────────────────────────────────────────────────────────────

def create_vbap(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VBAP"

    headers = ["VBELN", "POSNR", "MATNR", "ARKTX", "KWMENG", "VRKME", "NETPR", "MWSBP"]
    ws.append(headers)
    _style_header_row(ws, len(headers), fill_color="375623")

    for p in POSITIONS:
        ws.append([p[h] for h in headers])

    for col, width in zip("ABCDEFGH", [15, 10, 18, 35, 10, 8, 10, 8]):
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"  Created {path.name}")


# ─── orders.zip ───────────────────────────────────────────────────────────────

def create_orders_zip(path: Path) -> None:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for order in ORDERS:
            vbeln = order["VBELN"]
            for filename, title in order["files"]:
                zf.writestr(f"{vbeln}/{filename}", _fake_pdf(title))

    path.write_bytes(buf.getvalue())
    print(f"  Created {path.name}")
    with zipfile.ZipFile(path) as zf:
        for name in sorted(zf.namelist()):
            print(f"    {name}")


# ─── main ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("Generating test fixture files...")
    create_vbak(HERE / "VBAK.xlsx")
    create_vbap(HERE / "VBAP.xlsx")
    create_orders_zip(HERE / "orders.zip")
    print("Done.")
