"""Semantic file info system.

Provides ``FileInfo`` — a rich, type-aware view of any file's content.
Accessed via ``FileInput.get_info()``.  Parsers are lazy and cached.
"""

from __future__ import annotations

import base64
import io
import json
import logging
from typing import Any, Literal

from pydantic import BaseModel, ConfigDict

logger = logging.getLogger(__name__)

FileType = Literal["email", "pdf", "text", "spreadsheet", "structured", "image", "binary"]

# ---------------------------------------------------------------------------
# Extension → FileType classification
# ---------------------------------------------------------------------------

_EMAIL_EXTS = {"eml", "msg"}
_PDF_EXTS = {"pdf"}
_SPREADSHEET_EXTS = {"xlsx", "xls", "ods"}
_STRUCTURED_EXTS = {"json", "jsonl", "xml", "yaml", "yml", "toml"}
_IMAGE_EXTS = {"png", "jpg", "jpeg", "gif", "bmp", "svg", "webp", "tiff", "tif", "ico"}
_TEXT_EXTS = {
    # Plain text
    "txt", "text", "log", "ini", "cfg", "conf",
    # Data (non-structured — CSV/TSV are tabular text, not parsed into dicts)
    "csv", "tsv",
    # Web / docs
    "html", "htm", "xhtml", "md", "markdown", "rst", "tex",
    # Code
    "py", "js", "ts", "java", "c", "cpp", "h", "cs", "rb", "go", "rs",
    "sh", "bat", "ps1", "sql", "r", "swift", "kt", "scala", "php",
    "jsx", "tsx", "vue", "svelte", "css", "scss", "less", "sass",
}


# ---------------------------------------------------------------------------
# FileInfo model
# ---------------------------------------------------------------------------

class FileInfo(BaseModel):
    """Semantic view of a file's content.

    Every ``FileInfo`` has a universal ``.text`` property that returns
    extracted text for *any* file type.  Type-specific fields (e.g.
    ``.subject`` for emails, ``.pages`` for PDFs) are ``None`` when the
    type doesn't apply.
    """

    model_config = ConfigDict(populate_by_name=True)

    # --- Universal (always present) ----------------------------------------
    type: FileType
    filename: str
    size: int | None = None
    mime_type: str | None = None
    text: str

    # --- Email -------------------------------------------------------------
    subject: str | None = None
    sender: str | None = None
    recipients: list[str] | None = None
    to: list[str] | None = None
    cc: list[str] | None = None
    bcc: list[str] | None = None
    date: str | None = None
    body: str | None = None
    html: str | None = None
    headers: dict[str, str] | None = None
    attachments: list[FileInfo] | None = None

    # --- PDF ---------------------------------------------------------------
    pages: list[str] | None = None
    page_count: int | None = None
    metadata: dict[str, Any] | None = None

    # --- Text --------------------------------------------------------------
    encoding: str | None = None
    line_count: int | None = None

    # --- Spreadsheet -------------------------------------------------------
    sheet_names: list[str] | None = None
    sheets: dict[str, list[dict[str, Any]]] | None = None
    active_sheet: str | None = None
    rows: list[dict[str, Any]] | None = None
    columns: list[str] | None = None
    row_count: int | None = None

    # --- Structured (JSON / XML / YAML / TOML) -----------------------------
    data: Any | None = None
    format: str | None = None  # "json", "xml", "yaml", "toml"

    # --- Image -------------------------------------------------------------
    width: int | None = None
    height: int | None = None
    image_format: str | None = None  # "PNG", "JPEG", etc.
    mode: str | None = None  # "RGB", "RGBA", "L", etc.


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_file_info(file_input: Any) -> FileInfo:
    """Parse a ``FileInput`` into a ``FileInfo``.

    Determines file type from extension / MIME and dispatches to the
    appropriate parser.  Works regardless of whether the file arrived
    via ``path`` (post-resolution) or ``data`` (base64 from test runner).
    """
    ext = _ext(file_input.filename)
    mime = file_input.mime_type or ""
    file_type = _classify(ext, mime)
    raw = _get_bytes(file_input)

    parser = _PARSERS.get(file_type, _parse_binary)
    return parser(file_input, raw, ext)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _ext(filename: str) -> str:
    """Extract lowercase extension from a filename."""
    return filename.rsplit(".", 1)[-1].lower() if "." in filename else ""


def _classify(ext: str, mime: str) -> FileType:
    """Map extension + MIME to a FileType."""
    if ext in _EMAIL_EXTS or mime == "message/rfc822":
        return "email"
    if ext in _PDF_EXTS or mime == "application/pdf":
        return "pdf"
    if ext in _SPREADSHEET_EXTS:
        return "spreadsheet"
    if ext in _STRUCTURED_EXTS:
        return "structured"
    if ext in _IMAGE_EXTS or mime.startswith("image/"):
        return "image"
    if ext in _TEXT_EXTS or mime.startswith("text/"):
        return "text"
    return "binary"


def _get_bytes(file_input: Any) -> bytes | None:
    """Read raw bytes from whichever source is available."""
    # 1. Local path (post-resolution)
    if file_input.path:
        try:
            with open(file_input.path, "rb") as f:
                return f.read()
        except OSError:
            pass

    # 2. Base64 data (test runner)
    if file_input.data:
        try:
            return base64.b64decode(file_input.data)
        except Exception:
            pass

    return None


def _base_fields(file_input: Any) -> dict[str, Any]:
    """Common fields for every FileInfo."""
    return {
        "filename": file_input.filename,
        "size": file_input.size,
        "mime_type": file_input.mime_type,
    }


# ---------------------------------------------------------------------------
# Per-type parsers
# ---------------------------------------------------------------------------

def _parse_email(file_input: Any, raw: bytes | None, ext: str) -> FileInfo:
    if ext == "msg":
        return _parse_msg(file_input, raw)
    return _parse_eml(file_input, raw)


def _parse_eml(file_input: Any, raw: bytes | None) -> FileInfo:
    """Parse a .eml file with recursive flattening."""
    import email as _email
    from email.utils import getaddresses

    if raw is None:
        return FileInfo(type="email", text=f"[Email: {file_input.filename} — no data available]", **_base_fields(file_input))

    msg = _email.message_from_bytes(raw)

    subject = msg.get("Subject", "") or ""
    sender = msg.get("From", "") or ""
    date = msg.get("Date", "") or ""

    # Parse recipient lists
    to_raw = msg.get_all("To", [])
    cc_raw = msg.get_all("Cc", [])
    bcc_raw = msg.get_all("Bcc", [])
    to_list = [addr for _, addr in getaddresses(to_raw)] if to_raw else []
    cc_list = [addr for _, addr in getaddresses(cc_raw)] if cc_raw else []
    bcc_list = [addr for _, addr in getaddresses(bcc_raw)] if bcc_raw else []
    all_recipients = [r for r in to_list + cc_list + bcc_list if r]

    # Collect headers
    headers = {k: v for k, v in msg.items()}

    # Recursive walk — collect body text + attachments
    body_parts: list[str] = []
    html_parts: list[str] = []
    all_attachments: list[FileInfo] = []

    _walk_email(msg, body_parts, html_parts, all_attachments)

    body = "\n".join(body_parts).strip()
    html_body = "\n".join(html_parts).strip() or None

    # Build .text — formatted email like current _decode_eml
    lines = []
    if sender:
        lines.append(f"From: {sender}")
    if to_list:
        lines.append(f"To: {', '.join(to_list)}")
    if cc_list:
        lines.append(f"CC: {', '.join(cc_list)}")
    if date:
        lines.append(f"Date: {date}")
    if subject:
        lines.append(f"Subject: {subject}")
    lines.append("")
    lines.append(body)
    text = "\n".join(lines)

    return FileInfo(
        type="email",
        text=text,
        subject=subject or None,
        sender=sender or None,
        recipients=all_recipients or None,
        to=to_list or None,
        cc=cc_list or None,
        bcc=bcc_list or None,
        date=date or None,
        body=body or None,
        html=html_body,
        headers=headers or None,
        attachments=all_attachments or None,
        **_base_fields(file_input),
    )


def _walk_email(
    msg: Any,
    body_parts: list[str],
    html_parts: list[str],
    attachments: list[FileInfo],
) -> None:
    """Recursively walk MIME tree, collecting text + flattening attachments."""
    import email as _email

    if msg.is_multipart():
        for part in msg.get_payload():
            content_type = part.get_content_type()
            disposition = str(part.get("Content-Disposition", ""))

            if content_type == "message/rfc822":
                # Nested email — extract its body text, flatten its attachments
                nested_payload = part.get_payload()
                if isinstance(nested_payload, list):
                    for nested_msg in nested_payload:
                        nested_body: list[str] = []
                        nested_html: list[str] = []
                        _walk_email(nested_msg, nested_body, nested_html, attachments)
                        if nested_body:
                            subject = nested_msg.get("Subject", "")
                            sender = nested_msg.get("From", "")
                            body_parts.append(f"\n--- Forwarded message from {sender} ---")
                            if subject:
                                body_parts.append(f"Subject: {subject}")
                            body_parts.append("")
                            body_parts.extend(nested_body)
                elif nested_payload is not None:
                    nested_body_parts: list[str] = []
                    nested_html_parts: list[str] = []
                    _walk_email(nested_payload, nested_body_parts, nested_html_parts, attachments)
                    if nested_body_parts:
                        body_parts.extend(nested_body_parts)

            elif "attachment" in disposition:
                # File attachment — parse it into a FileInfo
                _extract_attachment(part, attachments)

            elif content_type == "text/plain":
                payload = part.get_payload(decode=True)
                if payload:
                    charset = part.get_content_charset() or "utf-8"
                    body_parts.append(payload.decode(charset, errors="replace"))

            elif content_type == "text/html":
                payload = part.get_payload(decode=True)
                if payload:
                    charset = part.get_content_charset() or "utf-8"
                    html_parts.append(payload.decode(charset, errors="replace"))

            else:
                # Inline content that isn't text — treat as attachment
                if part.get_payload(decode=True):
                    _extract_attachment(part, attachments)
    else:
        content_type = msg.get_content_type()
        payload = msg.get_payload(decode=True)
        if payload:
            charset = msg.get_content_charset() or "utf-8"
            if content_type == "text/plain":
                body_parts.append(payload.decode(charset, errors="replace"))
            elif content_type == "text/html":
                html_parts.append(payload.decode(charset, errors="replace"))
        elif content_type == "text/plain":
            text = msg.get_payload()
            if text:
                body_parts.append(text)


def _extract_attachment(part: Any, attachments: list[FileInfo]) -> None:
    """Extract a MIME part as a FileInfo and append to attachments."""
    payload = part.get_payload(decode=True)
    if not payload:
        return

    filename = part.get_filename() or "attachment"
    mime_type = part.get_content_type()
    ext = _ext(filename)
    file_type = _classify(ext, mime_type)

    # For attachments, parse based on their type
    parser = _PARSERS.get(file_type, _parse_binary)

    # Create a minimal fake file_input-like object for the parser
    class _AttachmentInput:
        def __init__(self, fn: str, data_bytes: bytes, mt: str, sz: int):
            self.filename = fn
            self.size = sz
            self.mime_type = mt
            self.path = None
            self.data = None
            self.extension = _ext(fn)
            self._raw = data_bytes

    att_input = _AttachmentInput(filename, payload, mime_type, len(payload))
    info = parser(att_input, payload, ext)
    attachments.append(info)


def _parse_msg(file_input: Any, raw: bytes | None) -> FileInfo:
    """Parse a .msg (Outlook) file."""
    if raw is None:
        return FileInfo(type="email", text=f"[Email: {file_input.filename} — no data available]", **_base_fields(file_input))

    try:
        import extract_msg  # type: ignore[import-untyped]
    except ImportError:
        return FileInfo(
            type="email",
            text=f"[Email: {file_input.filename} — install extract-msg to parse .msg files]",
            **_base_fields(file_input),
        )

    msg = extract_msg.Message(io.BytesIO(raw))
    subject = msg.subject or ""
    sender = msg.sender or ""
    to_str = msg.to or ""
    cc_str = msg.cc or ""
    date = msg.date or ""
    body = msg.body or ""

    to_list = [a.strip() for a in to_str.split(";") if a.strip()] if to_str else []
    cc_list = [a.strip() for a in cc_str.split(";") if a.strip()] if cc_str else []
    all_recipients = to_list + cc_list

    # Attachments
    all_attachments: list[FileInfo] = []
    for att in msg.attachments:
        att_data = att.data
        if att_data:
            att_filename = att.longFilename or att.shortFilename or "attachment"
            att_ext = _ext(att_filename)
            att_type = _classify(att_ext, "")

            class _MsgAtt:
                def __init__(self, fn: str, sz: int, mt: str):
                    self.filename = fn
                    self.size = sz
                    self.mime_type = mt
                    self.path = None
                    self.data = None
                    self.extension = _ext(fn)

            parser = _PARSERS.get(att_type, _parse_binary)
            info = parser(_MsgAtt(att_filename, len(att_data), ""), att_data, att_ext)
            all_attachments.append(info)

    lines = []
    if sender:
        lines.append(f"From: {sender}")
    if to_list:
        lines.append(f"To: {', '.join(to_list)}")
    if cc_list:
        lines.append(f"CC: {', '.join(cc_list)}")
    if date:
        lines.append(f"Date: {date}")
    if subject:
        lines.append(f"Subject: {subject}")
    lines.append("")
    lines.append(body)
    text = "\n".join(lines)

    msg.close()

    return FileInfo(
        type="email",
        text=text,
        subject=subject or None,
        sender=sender or None,
        recipients=all_recipients or None,
        to=to_list or None,
        cc=cc_list or None,
        date=str(date) if date else None,
        body=body or None,
        attachments=all_attachments or None,
        **_base_fields(file_input),
    )


# --- PDF -------------------------------------------------------------------

def _parse_pdf(file_input: Any, raw: bytes | None, ext: str) -> FileInfo:
    if raw is None:
        return FileInfo(type="pdf", text=f"[PDF: {file_input.filename} — no data available]", **_base_fields(file_input))

    try:
        from pypdf import PdfReader  # type: ignore[import-untyped]
    except ImportError:
        try:
            from PyPDF2 import PdfReader  # type: ignore[import-untyped]
        except ImportError:
            return FileInfo(
                type="pdf",
                text=f"[PDF: {file_input.filename} — install pypdf to extract text]",
                **_base_fields(file_input),
            )

    try:
        reader = PdfReader(io.BytesIO(raw))
    except Exception as e:
        return FileInfo(
            type="pdf",
            text=f"[PDF: {file_input.filename} — failed to parse: {e}]",
            **_base_fields(file_input),
        )

    pages = []
    for page in reader.pages:
        text = page.extract_text() or ""
        pages.append(text)

    meta = {}
    if reader.metadata:
        for key in ("/Title", "/Author", "/Subject", "/Creator", "/Producer"):
            val = reader.metadata.get(key)
            if val:
                meta[key.lstrip("/")] = str(val)

    full_text = "\n\n".join(pages)

    return FileInfo(
        type="pdf",
        text=full_text,
        pages=pages,
        page_count=len(pages),
        metadata=meta or None,
        **_base_fields(file_input),
    )


# --- Text ------------------------------------------------------------------

def _parse_text(file_input: Any, raw: bytes | None, ext: str) -> FileInfo:
    if raw is not None:
        # Try UTF-8 first, fall back to latin-1
        try:
            text = raw.decode("utf-8")
            enc = "utf-8"
        except UnicodeDecodeError:
            text = raw.decode("latin-1", errors="replace")
            enc = "latin-1"
    else:
        text = ""
        enc = None

    return FileInfo(
        type="text",
        text=text,
        encoding=enc,
        line_count=text.count("\n") + 1 if text else 0,
        **_base_fields(file_input),
    )


# --- Spreadsheet -----------------------------------------------------------

def _parse_spreadsheet(file_input: Any, raw: bytes | None, ext: str) -> FileInfo:
    if raw is None:
        return FileInfo(type="spreadsheet", text=f"[Spreadsheet: {file_input.filename} — no data available]", **_base_fields(file_input))

    try:
        import openpyxl  # type: ignore[import-untyped]
    except ImportError:
        return FileInfo(
            type="spreadsheet",
            text=f"[Spreadsheet: {file_input.filename} — install openpyxl to extract data]",
            **_base_fields(file_input),
        )

    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        return FileInfo(
            type="spreadsheet",
            text=f"[Spreadsheet: {file_input.filename} — failed to parse: {e}]",
            **_base_fields(file_input),
        )
    sheet_names = wb.sheetnames
    sheets: dict[str, list[dict[str, Any]]] = {}

    for name in sheet_names:
        ws = wb[name]
        rows_iter = ws.iter_rows(values_only=True)

        # First row = headers
        try:
            header_row = next(rows_iter)
        except StopIteration:
            sheets[name] = []
            continue

        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header_row)]
        sheet_rows: list[dict[str, Any]] = []
        for row in rows_iter:
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = val
            sheet_rows.append(row_dict)
        sheets[name] = sheet_rows

    wb.close()

    active = sheet_names[0] if sheet_names else None
    first_rows = sheets.get(active, []) if active else []
    first_cols = list(first_rows[0].keys()) if first_rows else []

    # Build CSV-like text from first sheet
    text_lines = []
    if first_cols:
        text_lines.append(",".join(str(c) for c in first_cols))
    for row in first_rows:
        text_lines.append(",".join(str(row.get(c, "")) for c in first_cols))
    text = "\n".join(text_lines)

    return FileInfo(
        type="spreadsheet",
        text=text,
        sheet_names=sheet_names,
        sheets=sheets,
        active_sheet=active,
        rows=first_rows,
        columns=first_cols,
        row_count=len(first_rows),
        **_base_fields(file_input),
    )


# --- Structured (JSON / XML / YAML / TOML) ---------------------------------

def _parse_structured(file_input: Any, raw: bytes | None, ext: str) -> FileInfo:
    if raw is None:
        return FileInfo(type="structured", text=f"[Structured: {file_input.filename} — no data available]", **_base_fields(file_input))

    text_content = raw.decode("utf-8", errors="replace")
    data: Any = None
    fmt: str | None = None

    if ext in ("json", "jsonl"):
        fmt = "json"
        try:
            if ext == "jsonl":
                data = [json.loads(line) for line in text_content.strip().splitlines() if line.strip()]
            else:
                data = json.loads(text_content)
        except json.JSONDecodeError:
            pass

    elif ext in ("yaml", "yml"):
        fmt = "yaml"
        try:
            import yaml  # type: ignore[import-untyped]
            data = yaml.safe_load(text_content)
        except ImportError:
            pass
        except Exception:
            pass

    elif ext == "toml":
        fmt = "toml"
        try:
            import tomllib
        except ImportError:
            try:
                import tomli as tomllib  # type: ignore[no-redef,import-untyped]
            except ImportError:
                tomllib = None  # type: ignore[assignment]
        if tomllib:
            try:
                data = tomllib.loads(text_content)
            except Exception:
                pass

    elif ext == "xml":
        fmt = "xml"
        try:
            import xml.etree.ElementTree as ET
            root = ET.fromstring(text_content)
            data = _xml_to_dict(root)
        except Exception:
            pass

    display_text = text_content
    if data is not None and fmt == "json":
        try:
            display_text = json.dumps(data, indent=2, ensure_ascii=False)
        except (TypeError, ValueError):
            pass

    return FileInfo(
        type="structured",
        text=display_text,
        data=data,
        format=fmt,
        **_base_fields(file_input),
    )


def _xml_to_dict(element: Any) -> dict[str, Any]:
    """Simple XML element → dict conversion."""
    result: dict[str, Any] = {}
    if element.attrib:
        result["@attributes"] = dict(element.attrib)
    children: dict[str, list[Any]] = {}
    for child in element:
        tag = child.tag
        child_dict = _xml_to_dict(child)
        children.setdefault(tag, []).append(child_dict)
    for tag, items in children.items():
        result[tag] = items[0] if len(items) == 1 else items
    if element.text and element.text.strip():
        if result:
            result["#text"] = element.text.strip()
        else:
            return element.text.strip()  # type: ignore[return-value]
    return result


# --- Image -----------------------------------------------------------------

def _parse_image(file_input: Any, raw: bytes | None, ext: str) -> FileInfo:
    if raw is None:
        return FileInfo(type="image", text=f"[Image: {file_input.filename}]", **_base_fields(file_input))

    width: int | None = None
    height: int | None = None
    img_format: str | None = None
    img_mode: str | None = None

    try:
        from PIL import Image  # type: ignore[import-untyped]
        img = Image.open(io.BytesIO(raw))
        width, height = img.size
        img_format = img.format
        img_mode = img.mode
        img.close()
    except ImportError:
        pass
    except Exception:
        pass

    parts = [f"Image: {file_input.filename}"]
    if width and height:
        parts.append(f"{width}x{height}")
    if img_format:
        parts.append(img_format)

    return FileInfo(
        type="image",
        text=f"[{', '.join(parts)}]",
        width=width,
        height=height,
        image_format=img_format,
        mode=img_mode,
        **_base_fields(file_input),
    )


# --- Binary (fallback) ----------------------------------------------------

def _parse_binary(file_input: Any, raw: bytes | None, ext: str) -> FileInfo:
    size = file_input.size
    if size is None and raw is not None:
        size = len(raw)
    mime = file_input.mime_type or "application/octet-stream"

    return FileInfo(
        type="binary",
        text=f"[Binary file: {file_input.filename}, {size or '?'} bytes, {mime}]",
        **_base_fields(file_input),
    )


# ---------------------------------------------------------------------------
# Parser dispatch table
# ---------------------------------------------------------------------------

_PARSERS: dict[FileType, Any] = {
    "email": _parse_email,
    "pdf": _parse_pdf,
    "text": _parse_text,
    "spreadsheet": _parse_spreadsheet,
    "structured": _parse_structured,
    "image": _parse_image,
    "binary": _parse_binary,
}
